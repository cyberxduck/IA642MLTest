from openpyxl import load_workbook
from openpyxl import Workbook
import csv

'''
	Coding by Alexander Veach
	IMPORTANT!
	This program was designed around the assumption of only safe ICMP then Only unsafe ICMP.
	If your dataset does not follow this rules it will not work.
	In fact it will ruin the dataset, so do not feed it anything other than Wireshark 
	ICMP captures as described in the Above & In the report.
'''

#Set's the file target, must be CSV
fileToClean = "betterCSV.csv"

#Declares Workbook
wb = Workbook()
ws = wb.active

#Transforms CSV to XLSX
with open(fileToClean, 'r') as f:
	for row in csv.reader(f):
		ws.append(row)

#Set length of dataset
wsRange = 502

#Pointer used in loop
iterator = 2

#Adds Label for Danger
ws['H1'].value = "Danger"

#Iterates through and adds a safe value up to original wsRange
while iterator < wsRange+1:
	ws['H' + str(iterator)].value = 'no'
	iterator += 1

#Set range to end of list
wsRange = 2815

#Iterates the rest of the list setting value to unsafe
while iterator < wsRange+1:
	ws['H' + str(iterator)].value = 'yes'
	iterator += 1

#Writes everything to a workable CSV
with open('veachTest.csv', 'w', newline="") as file_handle:
    csv_writer = csv.writer(file_handle)
    for row in ws.iter_rows():
        csv_writer.writerow([cell.value for cell in row])

